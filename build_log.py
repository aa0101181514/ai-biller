#!/usr/bin/env python3
"""
build_log.py — turn an aibiller CSV into a billing spreadsheet (.xlsx)

Reads aibiller_<project>[_<agent>].csv (written by aibiller.py) and emits an
Excel workbook billed on AI processing time:
  - AI time        = duration_min (OS-clock pure run time, start → stop)
  - billable hours = AI time rounded UP to a billing increment (default 15m)
  - tokens         = in/out + total (total excludes cache_read)

  Note: aibiller measures one OS-clock interval per segment; it cannot separate
  the user's reading/thinking from AI run time, so there is no separate
  "consultant input" column — AI time is the single, honest billing basis.

USAGE
    build_log.py --project NAME [--agent claude|codex]
    # claude and codex each produce their own .xlsx (agent in the filename)

CONFIG
    AIBILLER_HOME             fallback data dir (default: ~/.aibiller)
    AIBILLER_PROJECT_DIR      project root → <dir>/<project>_AI_BILLER/ (see aibiller.py)
    AIBILLER_BILLING_INCREMENT  billing round-up unit in hours (default: 0.25)
    AIBILLER_LANG             Excel language: en (default) | zh (Traditional Chinese)
                              (or pass --lang zh)

  Project data-dir resolution matches aibiller.py: a project init'd with
  `aibiller.py init` writes its Excel into the same <project>_AI_BILLER/ folder as
  its CSV. Otherwise it falls back to ~/.aibiller.

Requires: openpyxl  (pip install openpyxl)
MIT License.
"""
import csv
import math
import os
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def data_home(project=None):
    """Resolve the data dir for a project, sharing aibiller.py's logic so the
    Excel lands next to the CSV. Falls back to a standalone resolver if aibiller
    is not importable (keeps build_log usable on its own)."""
    try:
        import aibiller
        return aibiller._data_home(project)
    except Exception:
        if project:
            env_dir = os.environ.get("AIBILLER_PROJECT_DIR")
            if env_dir:
                return Path(env_dir) / f"{project}_AI_BILLER"
        return Path(os.environ.get("AIBILLER_HOME", str(Path.home() / ".aibiller")))


def billing_increment():
    try:
        return float(os.environ.get("AIBILLER_BILLING_INCREMENT", "0.25"))
    except ValueError:
        return 0.25


def lang(project=None):
    """Excel output language: 'en' (default) or 'zh' (Traditional Chinese).
    Priority: AIBILLER_LANG env > the project's registered lang (set by
    `aibiller.py init --lang`) > 'en'. 'zh'/'zh-tw'/'zh-hant' all map to zh."""
    try:
        import aibiller
        return aibiller._project_lang(project)
    except Exception:
        v = (os.environ.get("AIBILLER_LANG") or "en").lower()
        return "zh" if v.startswith("zh") else "en"


# i18n strings, keyed by language. headers() takes the increment for the dynamic
# round-up labels.
def _strings(inc):
    m = int(inc * 60)
    return {
        "en": {
            "sheet": "worklog",
            "headers": ["date", "start", "end", "task", "deliverable",
                        "AI time\n(min)",
                        "in_tok", "out_tok",
                        "total_tok\n(excl. cache_read)",
                        f"billable hours\n(AI time, round up {m}m)"],
            "title": "{project} — AI worklog{agent_tag} "
                     f"(AI processing time, billed by AI time, {m}m round-up)",
            "total": "TOTAL",
            "notes_sheet": "billing notes",
            "notes": [
                ("source", "aibiller CSV: {csv} (written by aibiller.py)"),
                ("AI time", "pure run time measured by aibiller via the OS clock (start → stop)"),
                ("billable hours", f"AI time rounded UP per {m}-minute unit"),
                ("tokens", "backfilled at stop from the agent transcript; total = in + out + cache_creation (cache_read excluded — background context, unrelated to work done)"),
                ("agent", "this sheet: {agent}; claude and codex keep separate CSVs/sheets, tokens matched per-transcript by time window"),
                ("update", "new segments enter the CSV via aibiller; re-run build_log.py --project {project}{agent_opt}"),
                ("total cost", "billable hours TOTAL × your agreed hourly rate"),
            ],
        },
        "zh": {
            "sheet": "工作計時log",
            "headers": ["日期", "開始", "結束", "工作項目", "產出／說明",
                        "AI處理時間\n(分)",
                        "in_tok", "out_tok",
                        "total_tok\n(不含cache_read)",
                        f"可計費工時\n(以AI處理時間×{m}分進位)"],
            "title": "{project} — AI 工作計時 log{agent_tag}"
                     f"（AI處理時間，以AI處理時間計費，{m}分進位）",
            "total": "合計",
            "notes_sheet": "計費口徑說明",
            "notes": [
                ("事實來源", "aibiller CSV：{csv}（由 aibiller.py start/stop 寫入）"),
                ("AI 處理時間", "aibiller 用 OS 時鐘量的純跑時間（start → stop）＝計費基礎"),
                ("可計費進位", f"以 AI 處理時間，每 {m} 分鐘無條件進位"),
                ("token", "stop 時掃對應 agent transcript 回填；total = in+out+cache_creation（不含 cache_read，後者為百萬級背景上下文與工作量無關）"),
                ("agent", "本表 agent = {agent}；claude 與 codex 各自獨立 CSV/Excel，token 按時間區間配對各自 transcript 不混"),
                ("更新方式", "新工作段以 aibiller 打點即自動進 CSV；重跑 build_log.py --project {project}{agent_opt} 更新本 Excel"),
                ("最終費用", "可計費工時合計 × 約定時薪"),
            ],
        },
    }


def csv_path(project, agent):
    home = data_home(project)
    if agent == "claude":
        return home / f"aibiller_{project}.csv"
    return home / f"aibiller_{project}_{agent}.csv"


def out_path(project, agent):
    suffix = "" if agent == "claude" else f"_{agent}"
    return data_home(project) / f"{project}_worklog{suffix}.xlsx"


def load_rows(project, agent):
    cp = csv_path(project, agent)
    if not cp.exists():
        return None
    with cp.open(encoding="utf-8") as f:
        return list(csv.DictReader(f))


def round_up(hours, inc):
    return math.ceil(hours / inc) * inc if hours > 0 else 0.0


def fnum(v, d=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return d


def fint(v, d=0):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return d


def build(project, agent):
    inc = billing_increment()
    rows = load_rows(project, agent)
    # init / first run: no CSV yet → still emit an empty-but-valid skeleton so the
    # folder is ready and the user can see the format. Real rows fill in on stop.
    if rows is None:
        rows = []

    lg = lang(project)
    L = _strings(inc)[lg]
    # Use a CJK-capable font for the zh sheet so WPS/Excel render cleanly.
    base_font_name = "Microsoft JhengHei" if lg == "zh" else None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = L["sheet"]

    title_font = Font(name=base_font_name, bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="155E5E")
    hdr_font = Font(name=base_font_name, bold=True, size=10, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    cell_font = Font(name=base_font_name, size=10)
    sum_font = Font(name=base_font_name, bold=True, size=11)
    sum_fill = PatternFill("solid", fgColor="FCE4D6")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center")

    headers = L["headers"]
    last_col = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    agent_tag = "" if agent == "claude" else (f"（{agent}）" if lg == "zh" else f" ({agent})")
    t = ws.cell(row=1, column=1,
                value=L["title"].format(project=project, agent_tag=agent_tag))
    t.font = title_font
    t.fill = title_fill
    t.alignment = center
    ws.row_dimensions[1].height = 28

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[2].height = 32

    r = 3
    tot_ai = tot_bill = 0.0
    tot_in = tot_out = tot_total = 0
    for row in rows:
        ws_iso = row.get("wall_start_iso", "")
        we_iso = row.get("wall_end_iso", "")
        start_hm = ws_iso[11:16] if len(ws_iso) >= 16 else ""
        end_hm = we_iso[11:16] if len(we_iso) >= 16 else ""
        ai_min = fnum(row.get("duration_min"))
        in_tok = fint(row.get("in_tok"))
        out_tok = fint(row.get("out_tok"))
        total_tok = fint(row.get("total_tok"))

        # Bill on AI processing time (OS-clock start→stop). aibiller cannot
        # separate the user's reading/thinking from AI run time, so there is no
        # honest "consultant input" column — AI time is the single billing basis.
        ai_h = ai_min / 60.0
        bill = round_up(ai_h, inc)
        tot_ai += ai_min
        tot_bill += bill
        tot_in += in_tok
        tot_out += out_tok
        tot_total += total_tok

        # task/deliverable: accept both the current English headers and the
        # earlier Chinese ones (工作項目/產出說明) so migrated CSVs still render.
        task = row.get("task") or row.get("工作項目") or ""
        deliverable = row.get("deliverable") or row.get("產出說明") or ""
        vals = [row.get("date", ""), start_hm, end_hm,
                task, deliverable,
                round(ai_min, 1),
                in_tok, out_tok, total_tok, bill]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = cell_font
            cell.border = border
            cell.alignment = wrap if c in (4, 5) else center
        r += 1

    ws.cell(row=r, column=4, value=L["total"]).font = sum_font
    for c in range(1, last_col + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = sum_fill
        cell.border = border
    for c, v in {6: round(tot_ai, 1),
                 7: tot_in, 8: tot_out, 9: tot_total,
                 10: round(tot_bill, 2)}.items():
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = sum_font
        cell.alignment = center

    widths = [12, 7, 7, 28, 46, 12, 10, 10, 14, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

    ws2 = wb.create_sheet(L["notes_sheet"])
    agent_opt = "" if agent == "claude" else f" --agent {agent}"
    fmt = {"csv": csv_path(project, agent).name, "agent": agent,
           "project": project, "agent_opt": agent_opt}
    for i, (k, v) in enumerate(L["notes"], 1):
        a = ws2.cell(row=i, column=1, value=k)
        a.font = Font(name=base_font_name, bold=True, size=10)
        b = ws2.cell(row=i, column=2, value=v.format(**fmt))
        b.font = Font(name=base_font_name, size=10)
        b.alignment = Alignment(wrap_text=True, vertical="top")
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 84

    out = out_path(project, agent)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"wrote: {out}")
    print(f"segments: {len(rows)} (agent={agent})")
    print(f"AI time: {tot_ai:.1f} min = {tot_ai/60:.2f} h")
    print(f"tokens: in={tot_in} out={tot_out} total(excl. read)={tot_total}")
    print(f"billable hours: {tot_bill:.2f} h (AI time, {int(inc*60)}m round-up)")
    return out


def main():
    raw = sys.argv[1:]
    project = None
    agent = "claude"
    if "--project" in raw:
        project = raw[raw.index("--project") + 1]
    if "--agent" in raw:
        agent = raw[raw.index("--agent") + 1]
    if "--lang" in raw:  # one-off override of AIBILLER_LANG
        os.environ["AIBILLER_LANG"] = raw[raw.index("--lang") + 1]
    if not project:
        print(__doc__)
        sys.exit(1)
    build(project, agent)


if __name__ == "__main__":
    main()
